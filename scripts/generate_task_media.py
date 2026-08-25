#!/usr/bin/env python3
"""
generate_task_media.py — build the four data files for bridge-survey-expense-claim.

Kept as a script rather than committing opaque binaries: the numbers in these
files are the task's ground truth, so they need to be reviewable and
regenerable.

    python3 scripts/generate_task_media.py bridge-survey-expense-claim/data

Produces:
    receipt_diner.jpg          the only place "two covers" and the struck-out
                               wine line exist -> the vision channel
    invoice_4482.pdf           two sites; only one is the claimant's
    site_budget_q3.xlsx        maps site -> project code -> lead surveyor
    reimbursement_policy.docx  meal cap, alcohol exclusion, own-project rule

Arithmetic the task resolves to:
    invoice, Bridge BR-114 only ... 148.00 + 96.50 + 23.80  = 268.30
    meals, 2 covers, capped ....... min(26.55, 2 x 12.00)   =  24.00
    claimable ..................................................  292.30
"""

import sys
import zipfile
from pathlib import Path

CLAIMABLE = 292.30


# --------------------------------------------------------------------- receipt
def build_receipt(path: Path) -> None:
    """Handwritten-style diner receipt. Two facts live only here."""
    from PIL import Image, ImageDraw, ImageFont

    W, H = 620, 880
    img = Image.new("RGB", (W, H), (247, 244, 236))
    d = ImageDraw.Draw(img)

    def font(sz, bold=False):
        for name in (("Helvetica-Bold" if bold else "Helvetica"),
                     "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold
                     else "/System/Library/Fonts/Supplemental/Arial.ttf"):
            try:
                return ImageFont.truetype(name, sz)
            except Exception:
                continue
        return ImageFont.load_default()

    d.text((60, 48), "COPPERLINE DINER", font=font(34, True), fill=(30, 30, 30))
    d.text((60, 96), "Route 9, Marchbanks", font=font(20), fill=(70, 70, 70))
    d.line((60, 140, W - 60, 140), fill=(120, 120, 120), width=2)

    # The covers count exists nowhere else in the task.
    d.text((60, 168), "TABLE 4          2 COVERS", font=font(26, True), fill=(25, 25, 25))
    d.text((60, 208), "14 Sept  19:42", font=font(20), fill=(70, 70, 70))
    d.line((60, 248, W - 60, 248), fill=(120, 120, 120), width=2)

    rows = [("Soup of the day", "8.50"),
            ("Steak pie", "11.25"),
            ("Coffee", "3.40"),
            ("Coffee", "3.40"),
            ("House red, glass", "9.00")]
    y = 286
    for i, (label, amt) in enumerate(rows):
        d.text((60, y), label, font=font(24), fill=(35, 35, 35))
        d.text((W - 150, y), amt, font=font(24), fill=(35, 35, 35))
        if i == len(rows) - 1:
            # struck through by hand -- the second fact that exists only here
            d.line((52, y + 17, W - 60, y + 15), fill=(190, 40, 40), width=3)
            d.text((60, y + 40), "(struck off - not on expenses)",
                   font=font(19), fill=(190, 40, 40))
        y += 52 if i < len(rows) - 1 else 92

    d.line((60, y + 8, W - 60, y + 8), fill=(120, 120, 120), width=2)
    d.text((60, y + 30), "TOTAL AS RUNG", font=font(26, True), fill=(25, 25, 25))
    d.text((W - 165, y + 30), "35.55", font=font(26, True), fill=(25, 25, 25))
    d.text((60, y + 78), "Server: Ines", font=font(19), fill=(90, 90, 90))

    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, "JPEG", quality=92)


# --------------------------------------------------------------------- invoice
def build_invoice(path: Path) -> None:
    """Single-page PDF, hand-assembled (no reportlab dependency)."""
    lines = [
        (72, 760, 16, "KESTREL SITE SERVICES LTD"),
        (72, 738, 10, "Unit 7, Halberd Way, Marchbanks"),
        (72, 712, 13, "INVOICE 4482          Issued 22 September 2026"),
        (72, 694, 10, "Bill to: Calloway Survey Partners - September site lot"),
        (72, 662, 12, "SITE: BRIDGE BR-114"),
        (86, 642, 11, "Laser level hire, 2 days ....................... 148.00"),
        (86, 624, 11, "Traffic marshal, half day ......................  96.50"),
        (86, 606, 11, "Consumables, marker pins .......................  23.80"),
        (72, 574, 12, "SITE: CULVERT CV-207"),
        (86, 554, 11, "GPR unit hire, 2 days .......................... 410.00"),
        (86, 536, 11, "Traffic marshal, full day ...................... 193.00"),
        (72, 496, 13, "INVOICE TOTAL ................................. 871.30"),
        (72, 462, 10, "Both sites appear on one invoice at the client's request."),
        (72, 446, 10, "Payment terms 30 days. Queries to accounts@kestrelsite.example."),
    ]
    content = "BT\n"
    for x, y, size, text in lines:
        esc = text.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")
        content += f"/F1 {size} Tf\n1 0 0 1 {x} {y} Tm\n({esc}) Tj\n"
    content += "ET\n"
    stream = content.encode("latin-1")

    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]

    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n".encode() + body + b"\nendobj\n"
    xref = len(out)
    out += f"xref\n0 {len(objs) + 1}\n".encode() + b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref}\n%%EOF\n").encode()

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(bytes(out))


# ----------------------------------------------------------------------- sheet
def build_budget(path: Path) -> None:
    """Project tracker: which site belongs to which code, and to whom."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Q3 Projects"
    headers = ["Project code", "Site", "Lead surveyor", "Q3 budget", "Q3 remaining"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for row in [
        ["SV-109", "Weir WR-88 resurvey", "D. Ferris", 3200.00, 410.00],
        ["SV-114", "Bridge BR-114", "R. Calloway", 4100.00, 1480.00],
        ["SV-121", "Embankment EM-31", "P. Oyelaran", 2750.00, 990.00],
        ["SV-207", "Culvert CV-207", "D. Ferris", 5400.00, 2050.00],
        ["SV-233", "Access road AR-12", "R. Calloway", 1900.00, 260.00],
    ]:
        ws.append(row)

    for col, width in zip("ABCDE", (14, 24, 18, 12, 14)):
        ws.column_dimensions[col].width = width

    note = ws.cell(row=8, column=1,
                   value="Claims must be filed against the code whose lead surveyor is the claimant.")
    note.font = Font(italic=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ------------------------------------------------------------------------- doc
def build_policy(path: Path) -> None:
    """Minimal but valid .docx (OOXML zip) — no python-docx dependency."""
    paras = [
        ("Calloway Survey Partners - Expense Reimbursement Policy", True),
        ("Revision 6, effective 1 September 2026. This revision supersedes any "
         "policy text held in the accounting system.", False),
        ("1. Project attribution", True),
        ("A claim may only include supplier lines for the project code whose lead "
         "surveyor is the claimant. Lines for any other code must be excluded from "
         "the claim, even where they appear on the same supplier invoice.", False),
        ("2. Subsistence", True),
        ("Meals are reimbursed at actual cost up to 12.00 per person per day. Where "
         "a receipt covers more than one person, the cap applies per cover. Amounts "
         "above the cap are not reimbursable and must not be claimed.", False),
        ("3. Alcohol", True),
        ("Alcoholic drinks are never reimbursable, whether or not they appear on an "
         "itemised receipt.", False),
        ("4. Filing", True),
        ("File one expense transaction per trip. Record the project code in the "
         "transaction description.", False),
    ]

    def xml_escape(t):
        return t.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    body = ""
    for text, bold in paras:
        rpr = "<w:rPr><w:b/></w:rPr>" if bold else ""
        body += (f'<w:p><w:r>{rpr}<w:t xml:space="preserve">'
                 f'{xml_escape(text)}</w:t></w:r></w:p>')

    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body>{body}<w:sectPr/></w:body></w:document>')

    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-'
        'officedocument.wordprocessingml.document.main+xml"/></Types>')

    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        'relationships/officeDocument" Target="word/document.xml"/></Relationships>')

    # python-docx expects the document part to have its own (possibly empty)
    # relationship part; omitting it makes some readers reject the package.
    doc_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        'relationships"/>')

    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document)
        z.writestr("word/_rels/document.xml.rels", doc_rels)


def main() -> int:
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "bridge-survey-expense-claim/data")
    build_receipt(out / "receipt_diner.jpg")
    build_invoice(out / "invoice_4482.pdf")
    build_budget(out / "site_budget_q3.xlsx")
    build_policy(out / "reimbursement_policy.docx")
    for f in sorted(out.iterdir()):
        print(f"  {f.name:28} {f.stat().st_size:>7} bytes")
    print(f"\n  claimable total encoded by these files: {CLAIMABLE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
